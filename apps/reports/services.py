"""Report dataset builders and multi-format exporters."""
from __future__ import annotations

import csv
import datetime
from io import BytesIO, StringIO

from django.http import HttpResponse
from django.utils import timezone

from apps.reports.models import ReportType


# ---------------------------------------------------------------------------
# Dataset builders
# ---------------------------------------------------------------------------
def build_dataset(report_type: str, params: dict | None = None) -> tuple[list[str], list[list], str]:
    """Return ``(headers, rows, title)`` for the requested report."""
    params = params or {}
    builder = {
        ReportType.EMPLOYEE: _employee_dataset,
        ReportType.DEPARTMENT: _department_dataset,
        ReportType.ATTENDANCE: _attendance_dataset,
        ReportType.PRODUCTIVITY: _productivity_dataset,
        ReportType.TIMESHEET: _timesheet_dataset,
        ReportType.PROJECT: _project_dataset,
        ReportType.PAYROLL: _payroll_dataset,
        ReportType.AUDIT: _audit_dataset,
    }.get(report_type, _employee_dataset)
    return builder(params)


def _parse_date(value, default):
    if not value:
        return default
    try:
        return datetime.datetime.strptime(value, "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return default


def _employee_dataset(params):
    from apps.employees.models import Employee

    headers = ["Code", "Name", "Email", "Department", "Job Title", "Status", "Hire Date"]
    rows = []
    for e in Employee.objects.select_related("user", "department", "job_title"):
        rows.append([
            e.employee_code,
            e.full_name,
            e.email,
            e.department.name if e.department else "",
            e.job_title.name if e.job_title else "",
            e.get_status_display(),
            e.hire_date.isoformat(),
        ])
    return headers, rows, "Employee Report"


def _department_dataset(params):
    from django.db.models import Count

    from apps.employees.models import Department

    headers = ["Department", "Code", "Manager", "Headcount", "Active"]
    rows = []
    for d in Department.objects.annotate(hc=Count("employees")).select_related("manager__user"):
        rows.append([
            d.name,
            d.code,
            d.manager.full_name if d.manager else "",
            d.hc,
            "Yes" if d.is_active else "No",
        ])
    return headers, rows, "Department Report"


def _attendance_dataset(params):
    from apps.attendance.models import AttendanceRecord

    end = timezone.localdate()
    start = end - datetime.timedelta(days=30)
    start = _parse_date(params.get("start"), start)
    end = _parse_date(params.get("end"), end)

    headers = ["Employee", "Date", "Status", "Clock In", "Clock Out", "Hours", "Late (min)"]
    rows = []
    qs = AttendanceRecord.objects.filter(
        date__gte=start, date__lte=end
    ).select_related("employee__user")
    for r in qs:
        rows.append([
            r.employee.full_name,
            r.date.isoformat(),
            r.get_status_display(),
            r.clock_in.strftime("%H:%M") if r.clock_in else "",
            r.clock_out.strftime("%H:%M") if r.clock_out else "",
            r.worked_hours,
            r.late_minutes,
        ])
    return headers, rows, f"Attendance Report ({start} to {end})"


def _productivity_dataset(params):
    from apps.productivity.models import ProductivityRecord

    end = timezone.localdate()
    start = end - datetime.timedelta(days=30)
    headers = ["Employee", "Date", "Productivity", "Activity", "Focus", "Active Hours", "Idle %"]
    rows = []
    qs = ProductivityRecord.objects.filter(
        date__gte=start, date__lte=end
    ).select_related("employee__user")
    for r in qs:
        rows.append([
            r.employee.full_name,
            r.date.isoformat(),
            r.productivity_score,
            r.activity_score,
            r.focus_score,
            r.active_hours,
            r.idle_percent,
        ])
    return headers, rows, "Productivity Report"


def _timesheet_dataset(params):
    from apps.timetracking.models import Timesheet

    headers = ["Employee", "Start", "End", "Total Hours", "Billable Hours", "Status"]
    rows = []
    for t in Timesheet.objects.select_related("employee__user"):
        rows.append([
            t.employee.full_name,
            t.start_date.isoformat(),
            t.end_date.isoformat(),
            t.total_hours,
            t.billable_hours,
            t.get_status_display(),
        ])
    return headers, rows, "Timesheet Report"


def _project_dataset(params):
    from apps.projects.models import Project

    headers = ["Project", "Code", "Client", "Status", "Progress", "Budget", "Spent", "Due"]
    rows = []
    for p in Project.objects.select_related("client"):
        rows.append([
            p.name,
            p.code,
            p.client.name if p.client else "",
            p.get_status_display(),
            f"{p.progress}%",
            p.budget,
            p.spent,
            p.due_date.isoformat() if p.due_date else "",
        ])
    return headers, rows, "Project Report"


def _payroll_dataset(params):
    from apps.payroll.models import Payslip

    headers = ["Reference", "Employee", "Period", "Gross", "Tax", "Deductions", "Net", "Status"]
    rows = []
    for p in Payslip.objects.select_related("employee__user", "period"):
        rows.append([
            p.reference,
            p.employee.full_name,
            p.period.name,
            p.gross_pay,
            p.tax_total,
            p.deduction_total,
            p.net_pay,
            p.get_status_display(),
        ])
    return headers, rows, "Payroll Report"


def _audit_dataset(params):
    from apps.audit.models import AuditLog

    headers = ["Timestamp", "Actor", "Action", "Module", "Target", "IP"]
    rows = []
    for log in AuditLog.objects.select_related("actor")[:5000]:
        rows.append([
            log.created_at.strftime("%Y-%m-%d %H:%M"),
            log.actor_repr,
            log.get_action_display(),
            log.module,
            log.target_repr,
            log.ip_address or "",
        ])
    return headers, rows, "Audit Report"


# ---------------------------------------------------------------------------
# Exporters
# ---------------------------------------------------------------------------
def export_csv(headers, rows, filename: str) -> HttpResponse:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)
    response = HttpResponse(buffer.getvalue(), content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
    return response


def export_xlsx(headers, rows, filename: str, title: str = "Report") -> HttpResponse:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return export_csv(headers, rows, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else "AA"].width = max(
            15, len(str(header)) + 2
        )

    buffer = BytesIO()
    wb.save(buffer)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
    return response


def export_pdf(headers, rows, filename: str, title: str = "Report") -> HttpResponse:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError:
        return export_csv(headers, rows, filename)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"]), Spacer(1, 12)]

    # Limit rows in PDF to keep the file reasonable.
    display_rows = rows[:500]
    data = [headers] + [[str(c) for c in row] for row in display_rows]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F4F6")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ])
    )
    elements.append(table)
    if len(rows) > len(display_rows):
        elements.append(Spacer(1, 8))
        elements.append(
            Paragraph(
                f"Showing {len(display_rows)} of {len(rows)} rows.", styles["Italic"]
            )
        )
    doc.build(elements)

    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}.pdf"'
    return response


def export(report_type: str, export_format: str, params: dict | None = None) -> HttpResponse:
    headers, rows, title = build_dataset(report_type, params)
    filename = f"{report_type}_{timezone.now():%Y%m%d_%H%M%S}"
    if export_format == "xlsx":
        return export_xlsx(headers, rows, filename, title)
    if export_format == "pdf":
        return export_pdf(headers, rows, filename, title)
    return export_csv(headers, rows, filename)
